import requests
import pandas as pd
import xml.etree.ElementTree as ET
import os
from datetime import datetime
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Font
import sys
import json
import xml.sax.saxutils as saxutils

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

PARAMS = {
    "URL": "",
    "CAMPOS": "",
    "FROM": "",
    "WHERE": "",
    "GROUP": "",
    "ORDER": ""
}

# =====================================================
# PARAMETROS
# =====================================================

print("JSON recebido:", sys.argv[1])

try:

    with open(sys.argv[1], "r", encoding="utf-8-sig") as f:
        PARAMS = json.load(f)

    print("Parametros:", PARAMS)

except Exception as e:

    print("Erro ao interpretar JSON:", str(e))
    sys.exit(1)

URL   = PARAMS.get("URL","")
CAMPOS= PARAMS.get("CAMPOS","")
FROM  = PARAMS.get("FROM","")
WHERE = PARAMS.get("WHERE","")
GROUP = PARAMS.get("GROUP","")
ORDER = PARAMS.get("ORDER","")

usuario = os.getenv("PY_USER")
senha   = os.getenv("PY_PASS")

if not usuario:
    print("Usuario nao informado")
    sys.exit(1)

if not senha:
    print("Senha nao informada")
    sys.exit(1)

if not URL:
    print("Parametro URL nao informado")
    sys.exit(1)

HEADERS = {
    "Content-Type": "text/xml; charset=utf-8",
    "SOAPAction": "ExportaDados"
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
arquivo_excel = os.path.join(BASE_DIR, f"ExportaDados_{timestamp}.xlsx")

# =====================================================
# FUNCAO ESCAPE XML
# =====================================================

def xml_escape(text):
    if not text:
        return ""
    return saxutils.escape(text)

# =====================================================
# MONTAR SOAP
# =====================================================

def montar_soap_body():

    campos = xml_escape(CAMPOS)
    from_  = xml_escape(FROM)
    where  = xml_escape(WHERE)
    group_ = xml_escape(GROUP)
    order_ = xml_escape(ORDER)

    return f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                  xmlns:ser="http://services.senior.com.br">
   <soapenv:Header/>
   <soapenv:Body>
      <ser:ExportaDados>
         <user>{usuario}</user>
         <password>{senha}</password>
         <encryption>0</encryption>
         <parameters>
            <Campos>{campos}</Campos>
            <From>{from_}</From>
            <Where>{where}</Where>
            <Group>{group_}</Group>
            <Order>{order_}</Order>
         </parameters>
      </ser:ExportaDados>
   </soapenv:Body>
</soapenv:Envelope>
"""

# =====================================================
# REQUISIÇÃO
# =====================================================

print("Enviando requisição SOAP...")

soap = montar_soap_body()

print("\nSOAP enviado (inicio):\n")
print(soap[:1000])
print("\n---------------------------------\n")

try:

    response = requests.post(
        URL,
        data=soap.encode("utf-8"),
        headers=HEADERS,
        verify=False,
        timeout=120
    )

except Exception as e:

    print("Erro de conexao:", str(e))
    sys.exit(1)

if response.status_code != 200:

    print("\nErro HTTP:", response.status_code)
    print("\nResposta completa:\n")
    print(response.text)
    sys.exit(1)

print("Resposta recebida.")

# =====================================================
# PARSE XML
# =====================================================

try:
    root = ET.fromstring(response.content)

except Exception as e:

    print("Erro ao interpretar XML:", str(e))
    sys.exit(1)

result_node = None

for elem in root.iter():

    if elem.tag.split("}")[-1].lower() == "result":
        result_node = elem
        break

if result_node is None:

    print("Tag <result> não encontrada")
    print(response.text[:500])
    sys.exit(1)

# =====================================================
# VERIFICAR ERRO DO SERVIÇO
# =====================================================

for elem in result_node.iter():

    tag = elem.tag.split("}")[-1].lower()

    if tag == "mensagemretorno" and elem.text:

        msg = elem.text.strip()
        print(msg)

        if "erro" in msg.lower():
            sys.exit(1)

# =====================================================
# LOCALIZAR JSON
# =====================================================

json_text = ""

for elem in result_node.iter():

    tag = elem.tag.split("}")[-1].lower()

    if tag == "json" and elem.text:

        json_text += elem.text.strip() + "\n"

if not json_text:

    print("Tag <json> não encontrada")
    print(response.text[:500])
    sys.exit(1)

print("Primeiros caracteres do JSON:", json_text[:200])

# =====================================================
# CONVERTER JSON
# =====================================================

dados = []

try:

    for linha in json_text.splitlines():

        linha = linha.strip()

        if linha:
            dados.append(json.loads(linha))

except Exception as e:

    print("Erro ao converter JSON:", str(e))
    print("Conteudo recebido:", json_text[:500])
    sys.exit(1)

print("OK Quantidade de registros:", len(dados))

# =====================================================
# GERAR EXCEL
# =====================================================

try:

    if not dados:

        print("Nenhum registro retornado")
        sys.exit(0)

    print("Criando DataFrame...")

    df = pd.DataFrame(dados)

    print("Removendo linhas vazias...")

    df.dropna(how="all", inplace=True)

    print("Gerando Excel:", arquivo_excel)

    df.to_excel(arquivo_excel, index=False)

    print("Excel gravado")

except Exception as e:

    print("Erro ao gerar Excel:", str(e))
    sys.exit(1)

# =====================================================
# FORMATAR EXCEL
# =====================================================

try:

    print("Formatando Excel...")

    wb = load_workbook(arquivo_excel)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:

        max_length = 0
        column = col[0].column_letter

        for cell in col:

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column].width = max_length + 2

    wb.save(arquivo_excel)

    print("Formatacao concluida")

except Exception as e:

    print("Erro ao formatar Excel:", str(e))

# =====================================================
# RESULTADO
# =====================================================

print("Excel gerado com sucesso")
print("Arquivo:", arquivo_excel)
print("Total de registros:", len(df))